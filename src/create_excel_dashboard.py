"""YES24 IT 베스트셀러 CSV 데이터를 분석하여 엑셀 대시보드를 생성한다."""

import os
import re
from collections import Counter

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

# ── 경로 설정 ────────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_PATH = os.path.join(BASE, "data", "yes24_it_bestseller.csv")
OUTPUT_PATH = os.path.join(BASE, "data", "yes24_it_bestseller_dashboard.xlsx")

# ── 스타일 상수 ──────────────────────────────────────────────────────────────
DARK_BG = "1B2A4A"
HEADER_BG = "2D3A5C"
ACCENT = "6C63FF"
LIGHT_BG = "E8EAF6"
WHITE = "FFFFFF"
DARK_TEXT = "1B2A4A"
LIGHT_TEXT = "E8EAF6"
GRAY_TEXT = "9E9E9E"

HEADER_FONT = Font(name="Arial", bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill("solid", fgColor=HEADER_BG)
TITLE_FONT = Font(name="Arial", bold=True, color=ACCENT, size=16)
SUBTITLE_FONT = Font(name="Arial", bold=True, color=DARK_TEXT, size=13)
METRIC_FONT = Font(name="Arial", bold=True, color=ACCENT, size=22)
METRIC_LABEL_FONT = Font(name="Arial", color=GRAY_TEXT, size=10)
DATA_FONT = Font(name="Arial", color=DARK_TEXT, size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def load_and_preprocess() -> pd.DataFrame:
    df = pd.read_csv(DATA_PATH)
    df["판매가(원)"] = pd.to_numeric(df["판매가(원)"].astype(str).str.replace(",", ""), errors="coerce")
    df["평점"] = pd.to_numeric(df["평점"], errors="coerce")
    df["리뷰수"] = pd.to_numeric(df["리뷰수"].astype(str).str.replace(",", ""), errors="coerce").fillna(0).astype(int)
    df["출판년도"] = df["출판일"].astype(str).str.extract(r"(\d{4})년")
    df["출판월"] = df["출판일"].astype(str).str.extract(r"(\d{2})월")
    df["출판년월"] = df["출판년도"] + "-" + df["출판월"]
    return df


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_area(ws, start_row, end_row, max_col):
    stripe = PatternFill("solid", fgColor=LIGHT_BG)
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER if c != 2 else LEFT
            if (r - start_row) % 2 == 1:
                cell.fill = stripe


def set_col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def write_metric_card(ws, row, col, label, value):
    ws.cell(row=row, column=col, value=label).font = METRIC_LABEL_FONT
    ws.cell(row=row, column=col).alignment = CENTER
    ws.cell(row=row + 1, column=col, value=value).font = METRIC_FONT
    ws.cell(row=row + 1, column=col).alignment = CENTER


# ══════════════════════════════════════════════════════════════════════════════
#  시트 1: 대시보드
# ══════════════════════════════════════════════════════════════════════════════
def create_dashboard(wb: Workbook, df: pd.DataFrame):
    ws = wb.active
    ws.title = "대시보드"
    ws.sheet_properties.tabColor = ACCENT

    # 배경색
    bg = PatternFill("solid", fgColor=WHITE)

    # 타이틀
    ws.merge_cells("A1:H1")
    ws.cell(row=1, column=1, value="YES24 IT 베스트셀러 대시보드").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    ws.cell(row=2, column=1, value=f"총 {len(df):,}권 | 평균 평점 {df['평점'].mean():.2f} | 평균 가격 {int(df['판매가(원)'].mean()):,}원 | 총 리뷰 {df['리뷰수'].sum():,}건").font = Font(name="Arial", color=GRAY_TEXT, size=11)
    ws.cell(row=2, column=1).alignment = CENTER
    ws.row_dimensions[2].height = 25

    # ── 핵심 지표 카드 (row 4-5) ────────────────────────────────────────────
    write_metric_card(ws, 4, 1, "총 도서 수", f"{len(df):,}권")
    write_metric_card(ws, 4, 3, "평균 평점", f"{df['평점'].mean():.2f}")
    write_metric_card(ws, 4, 5, "중앙값 가격", f"{int(df['판매가(원)'].median()):,}원")
    write_metric_card(ws, 4, 7, "총 리뷰", f"{df['리뷰수'].sum():,}건")
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 35

    # ── 출판사 Top 10 (row 8~) ─────────────────────────────────────────────
    ws.cell(row=7, column=1, value="출판사별 베스트셀러 수 (Top 10)").font = SUBTITLE_FONT
    pub_counts = df["출판사"].value_counts().head(10).reset_index()
    pub_counts.columns = ["출판사", "도서 수"]
    start_row = 8
    ws.cell(row=start_row, column=1, value="출판사")
    ws.cell(row=start_row, column=2, value="도서 수")
    style_header_row(ws, start_row, 2)
    for i, (_, r) in enumerate(pub_counts.iterrows()):
        ws.cell(row=start_row + 1 + i, column=1, value=r["출판사"])
        ws.cell(row=start_row + 1 + i, column=2, value=int(r["도서 수"]))
    end_row = start_row + len(pub_counts)
    style_data_area(ws, start_row + 1, end_row, 2)

    chart_pub = BarChart()
    chart_pub.type = "bar"
    chart_pub.title = "출판사별 베스트셀러 수"
    chart_pub.x_axis.title = "도서 수"
    chart_pub.y_axis.title = None
    chart_pub.style = 10
    chart_pub.width = 22
    chart_pub.height = 12
    cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=end_row)
    vals = Reference(ws, min_col=2, min_row=start_row, max_row=end_row)
    chart_pub.add_data(vals, titles_from_data=True)
    chart_pub.set_categories(cats)
    chart_pub.shape = 4
    s = chart_pub.series[0]
    s.graphicalProperties.solidFill = ACCENT
    ws.add_chart(chart_pub, "D7")

    # ── 저자 Top 10 ────────────────────────────────────────────────────────
    auth_start = end_row + 3
    ws.cell(row=auth_start - 1, column=1, value="저자별 베스트셀러 수 (Top 10)").font = SUBTITLE_FONT
    auth_counts = df["저자"].value_counts().head(10).reset_index()
    auth_counts.columns = ["저자", "도서 수"]
    ws.cell(row=auth_start, column=1, value="저자")
    ws.cell(row=auth_start, column=2, value="도서 수")
    style_header_row(ws, auth_start, 2)
    for i, (_, r) in enumerate(auth_counts.iterrows()):
        ws.cell(row=auth_start + 1 + i, column=1, value=str(r["저자"]))
        ws.cell(row=auth_start + 1 + i, column=2, value=int(r["도서 수"]))
    auth_end = auth_start + len(auth_counts)
    style_data_area(ws, auth_start + 1, auth_end, 2)

    chart_auth = BarChart()
    chart_auth.type = "bar"
    chart_auth.title = "저자별 베스트셀러 수"
    chart_auth.style = 10
    chart_auth.width = 22
    chart_auth.height = 12
    cats2 = Reference(ws, min_col=1, min_row=auth_start + 1, max_row=auth_end)
    vals2 = Reference(ws, min_col=2, min_row=auth_start, max_row=auth_end)
    chart_auth.add_data(vals2, titles_from_data=True)
    chart_auth.set_categories(cats2)
    s2 = chart_auth.series[0]
    s2.graphicalProperties.solidFill = "FF6F61"
    ws.add_chart(chart_auth, f"D{auth_start - 1}")

    set_col_widths(ws, {"A": 22, "B": 12, "C": 3, "D": 18, "E": 18, "F": 3, "G": 18, "H": 18})


# ══════════════════════════════════════════════════════════════════════════════
#  시트 2: 가격/평점 분석
# ══════════════════════════════════════════════════════════════════════════════
def create_price_rating_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("가격·평점 분석")
    ws.sheet_properties.tabColor = "FF6F61"

    ws.merge_cells("A1:H1")
    ws.cell(row=1, column=1, value="가격 & 평점 분석").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = CENTER
    ws.row_dimensions[1].height = 35

    # ── 가격대별 분포 (10,000원 단위) ──────────────────────────────────────
    ws.cell(row=3, column=1, value="가격대별 도서 수").font = SUBTITLE_FONT
    bins = list(range(0, int(df["판매가(원)"].max()) + 10000, 10000))
    labels = [f"{b//10000}만원대" for b in bins[:-1]]
    df_temp = df.copy()
    df_temp["가격대"] = pd.cut(df_temp["판매가(원)"], bins=bins, labels=labels, right=False)
    price_dist = df_temp["가격대"].value_counts().sort_index().reset_index()
    price_dist.columns = ["가격대", "도서 수"]

    start = 4
    ws.cell(row=start, column=1, value="가격대")
    ws.cell(row=start, column=2, value="도서 수")
    style_header_row(ws, start, 2)
    for i, (_, r) in enumerate(price_dist.iterrows()):
        ws.cell(row=start + 1 + i, column=1, value=str(r["가격대"]))
        ws.cell(row=start + 1 + i, column=2, value=int(r["도서 수"]))
    end = start + len(price_dist)
    style_data_area(ws, start + 1, end, 2)

    chart_price = BarChart()
    chart_price.title = "가격대별 도서 수 분포"
    chart_price.style = 10
    chart_price.width = 22
    chart_price.height = 12
    cats = Reference(ws, min_col=1, min_row=start + 1, max_row=end)
    vals = Reference(ws, min_col=2, min_row=start, max_row=end)
    chart_price.add_data(vals, titles_from_data=True)
    chart_price.set_categories(cats)
    chart_price.series[0].graphicalProperties.solidFill = "4CAF50"
    ws.add_chart(chart_price, "D3")

    # ── 평점 구간별 분포 ──────────────────────────────────────────────────
    rating_start = end + 3
    ws.cell(row=rating_start - 1, column=1, value="평점 구간별 도서 수").font = SUBTITLE_FONT
    rating_bins = [0, 5, 7, 8, 9, 9.5, 10.01]
    rating_labels = ["~5.0", "5.1~7.0", "7.1~8.0", "8.1~9.0", "9.1~9.5", "9.6~10.0"]
    df_r = df[df["평점"].notna()].copy()
    df_r["평점구간"] = pd.cut(df_r["평점"], bins=rating_bins, labels=rating_labels)
    rating_dist = df_r["평점구간"].value_counts().sort_index().reset_index()
    rating_dist.columns = ["평점 구간", "도서 수"]

    ws.cell(row=rating_start, column=1, value="평점 구간")
    ws.cell(row=rating_start, column=2, value="도서 수")
    style_header_row(ws, rating_start, 2)
    for i, (_, r) in enumerate(rating_dist.iterrows()):
        ws.cell(row=rating_start + 1 + i, column=1, value=str(r["평점 구간"]))
        ws.cell(row=rating_start + 1 + i, column=2, value=int(r["도서 수"]))
    r_end = rating_start + len(rating_dist)
    style_data_area(ws, rating_start + 1, r_end, 2)

    chart_rating = BarChart()
    chart_rating.title = "평점 구간별 도서 수"
    chart_rating.style = 10
    chart_rating.width = 22
    chart_rating.height = 12
    cats_r = Reference(ws, min_col=1, min_row=rating_start + 1, max_row=r_end)
    vals_r = Reference(ws, min_col=2, min_row=rating_start, max_row=r_end)
    chart_rating.add_data(vals_r, titles_from_data=True)
    chart_rating.set_categories(cats_r)
    chart_rating.series[0].graphicalProperties.solidFill = "9C27B0"
    ws.add_chart(chart_rating, f"D{rating_start - 1}")

    # ── 기술 통계 요약 ────────────────────────────────────────────────────
    stat_start = r_end + 3
    ws.cell(row=stat_start - 1, column=1, value="기술 통계 요약").font = SUBTITLE_FONT
    desc = df[["판매가(원)", "평점", "리뷰수"]].describe().round(2)
    ws.cell(row=stat_start, column=1, value="통계 항목")
    ws.cell(row=stat_start, column=2, value="판매가(원)")
    ws.cell(row=stat_start, column=3, value="평점")
    ws.cell(row=stat_start, column=4, value="리뷰수")
    style_header_row(ws, stat_start, 4)
    for i, (idx, row_data) in enumerate(desc.iterrows()):
        ws.cell(row=stat_start + 1 + i, column=1, value=idx)
        ws.cell(row=stat_start + 1 + i, column=2, value=row_data["판매가(원)"])
        ws.cell(row=stat_start + 1 + i, column=3, value=row_data["평점"] if pd.notna(row_data["평점"]) else "")
        ws.cell(row=stat_start + 1 + i, column=4, value=int(row_data["리뷰수"]))
    stat_end = stat_start + len(desc)
    style_data_area(ws, stat_start + 1, stat_end, 4)

    set_col_widths(ws, {"A": 18, "B": 14, "C": 14, "D": 14, "E": 3, "F": 18, "G": 18, "H": 18})


# ══════════════════════════════════════════════════════════════════════════════
#  시트 3: 출판 추이
# ══════════════════════════════════════════════════════════════════════════════
def create_trend_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("출판 추이")
    ws.sheet_properties.tabColor = "4CAF50"

    ws.merge_cells("A1:H1")
    ws.cell(row=1, column=1, value="출판 추이 분석").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = CENTER
    ws.row_dimensions[1].height = 35

    # ── 연도별 추이 ────────────────────────────────────────────────────────
    ws.cell(row=3, column=1, value="연도별 베스트셀러 출시 수").font = SUBTITLE_FONT
    year_dist = df["출판년도"].value_counts().sort_index().reset_index()
    year_dist.columns = ["출판년도", "도서 수"]

    start = 4
    ws.cell(row=start, column=1, value="출판년도")
    ws.cell(row=start, column=2, value="도서 수")
    style_header_row(ws, start, 2)
    for i, (_, r) in enumerate(year_dist.iterrows()):
        ws.cell(row=start + 1 + i, column=1, value=str(r["출판년도"]))
        ws.cell(row=start + 1 + i, column=2, value=int(r["도서 수"]))
    end = start + len(year_dist)
    style_data_area(ws, start + 1, end, 2)

    chart_year = LineChart()
    chart_year.title = "연도별 베스트셀러 출시 추이"
    chart_year.style = 10
    chart_year.width = 24
    chart_year.height = 14
    chart_year.y_axis.title = "도서 수"
    cats = Reference(ws, min_col=1, min_row=start + 1, max_row=end)
    vals = Reference(ws, min_col=2, min_row=start, max_row=end)
    chart_year.add_data(vals, titles_from_data=True)
    chart_year.set_categories(cats)
    chart_year.series[0].graphicalProperties.line.solidFill = ACCENT
    chart_year.series[0].graphicalProperties.line.width = 25000
    ws.add_chart(chart_year, "D3")

    # ── 월별 추이 (최근 3년) ──────────────────────────────────────────────
    month_start = end + 3
    ws.cell(row=month_start - 1, column=1, value="월별 추이 (최근 3년)").font = SUBTITLE_FONT
    recent = df[df["출판년도"].astype(str) >= "2023"].copy()
    month_dist = recent.groupby("출판년월").size().reset_index(name="도서 수").sort_values("출판년월")

    ws.cell(row=month_start, column=1, value="출판년월")
    ws.cell(row=month_start, column=2, value="도서 수")
    style_header_row(ws, month_start, 2)
    for i, (_, r) in enumerate(month_dist.iterrows()):
        ws.cell(row=month_start + 1 + i, column=1, value=str(r["출판년월"]))
        ws.cell(row=month_start + 1 + i, column=2, value=int(r["도서 수"]))
    m_end = month_start + len(month_dist)
    style_data_area(ws, month_start + 1, m_end, 2)

    chart_month = LineChart()
    chart_month.title = "월별 베스트셀러 추이 (2023~)"
    chart_month.style = 10
    chart_month.width = 24
    chart_month.height = 14
    chart_month.y_axis.title = "도서 수"
    cats_m = Reference(ws, min_col=1, min_row=month_start + 1, max_row=m_end)
    vals_m = Reference(ws, min_col=2, min_row=month_start, max_row=m_end)
    chart_month.add_data(vals_m, titles_from_data=True)
    chart_month.set_categories(cats_m)
    chart_month.series[0].graphicalProperties.line.solidFill = "FF6F61"
    chart_month.series[0].graphicalProperties.line.width = 25000
    chart_month.x_axis.tickLblPos = "low"
    ws.add_chart(chart_month, f"D{month_start - 1}")

    set_col_widths(ws, {"A": 16, "B": 12, "C": 3, "D": 18, "E": 18, "F": 3, "G": 18, "H": 18})


# ══════════════════════════════════════════════════════════════════════════════
#  시트 4: 키워드 분석
# ══════════════════════════════════════════════════════════════════════════════
def create_keyword_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("키워드 분석")
    ws.sheet_properties.tabColor = "FF9800"

    ws.merge_cells("A1:H1")
    ws.cell(row=1, column=1, value="도서 제목 키워드 분석").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = CENTER
    ws.row_dimensions[1].height = 35

    stopwords = {"의", "에", "를", "이", "은", "는", "가", "와", "과", "도", "로", "으로",
                 "for", "the", "and", "with", "to", "a", "an", "in", "on", "of"}
    words = []
    for t in df["도서명"].dropna():
        tokens = re.findall(r"[가-힣]{2,}|[A-Za-z]{3,}", str(t))
        words.extend(w for w in tokens if w.lower() not in stopwords)
    kw = Counter(words).most_common(20)

    ws.cell(row=3, column=1, value="상위 20개 키워드").font = SUBTITLE_FONT
    start = 4
    ws.cell(row=start, column=1, value="키워드")
    ws.cell(row=start, column=2, value="빈도")
    style_header_row(ws, start, 2)
    for i, (word, count) in enumerate(kw):
        ws.cell(row=start + 1 + i, column=1, value=word)
        ws.cell(row=start + 1 + i, column=2, value=count)
    end = start + len(kw)
    style_data_area(ws, start + 1, end, 2)

    chart_kw = BarChart()
    chart_kw.type = "bar"
    chart_kw.title = "도서 제목 주요 키워드 (Top 20)"
    chart_kw.style = 10
    chart_kw.width = 24
    chart_kw.height = 14
    cats = Reference(ws, min_col=1, min_row=start + 1, max_row=end)
    vals = Reference(ws, min_col=2, min_row=start, max_row=end)
    chart_kw.add_data(vals, titles_from_data=True)
    chart_kw.set_categories(cats)
    chart_kw.series[0].graphicalProperties.solidFill = "FF9800"
    ws.add_chart(chart_kw, "D3")

    # ── 출판사별 평균 가격 ────────────────────────────────────────────────
    pub_price_start = end + 3
    ws.cell(row=pub_price_start - 1, column=1, value="출판사별 평균 가격 (Top 10)").font = SUBTITLE_FONT
    pub_price = df.groupby("출판사")["판매가(원)"].mean().sort_values(ascending=False).head(10).reset_index()
    pub_price.columns = ["출판사", "평균 가격(원)"]
    pub_price["평균 가격(원)"] = pub_price["평균 가격(원)"].round(0).astype(int)

    ws.cell(row=pub_price_start, column=1, value="출판사")
    ws.cell(row=pub_price_start, column=2, value="평균 가격(원)")
    style_header_row(ws, pub_price_start, 2)
    for i, (_, r) in enumerate(pub_price.iterrows()):
        ws.cell(row=pub_price_start + 1 + i, column=1, value=r["출판사"])
        ws.cell(row=pub_price_start + 1 + i, column=2, value=r["평균 가격(원)"])
    pp_end = pub_price_start + len(pub_price)
    style_data_area(ws, pub_price_start + 1, pp_end, 2)

    chart_pp = BarChart()
    chart_pp.type = "bar"
    chart_pp.title = "출판사별 평균 가격"
    chart_pp.style = 10
    chart_pp.width = 24
    chart_pp.height = 14
    cats_pp = Reference(ws, min_col=1, min_row=pub_price_start + 1, max_row=pp_end)
    vals_pp = Reference(ws, min_col=2, min_row=pub_price_start, max_row=pp_end)
    chart_pp.add_data(vals_pp, titles_from_data=True)
    chart_pp.set_categories(cats_pp)
    chart_pp.series[0].graphicalProperties.solidFill = "E91E63"
    ws.add_chart(chart_pp, f"D{pub_price_start - 1}")

    set_col_widths(ws, {"A": 18, "B": 16, "C": 3, "D": 18, "E": 18, "F": 3, "G": 18, "H": 18})


# ══════════════════════════════════════════════════════════════════════════════
#  시트 5: 원본 데이터
# ══════════════════════════════════════════════════════════════════════════════
def create_data_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("데이터")
    ws.sheet_properties.tabColor = "607D8B"

    cols = ["순위", "도서명", "저자", "출판사", "출판일", "판매가(원)", "평점", "리뷰수", "상세링크"]
    for c, col_name in enumerate(cols, 1):
        ws.cell(row=1, column=c, value=col_name)
    style_header_row(ws, 1, len(cols))

    for r, (_, row_data) in enumerate(df.iterrows(), 2):
        for c, col_name in enumerate(cols, 1):
            val = row_data[col_name]
            ws.cell(row=r, column=c, value=val if pd.notna(val) else "")

    end_row = len(df) + 1
    style_data_area(ws, 2, end_row, len(cols))

    # 필터
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{end_row}"

    set_col_widths(ws, {"A": 8, "B": 45, "C": 22, "D": 18, "E": 14, "F": 14, "G": 8, "H": 10, "I": 40})


# ══════════════════════════════════════════════════════════════════════════════
#  메인 실행
# ══════════════════════════════════════════════════════════════════════════════
def main():
    df = load_and_preprocess()

    wb = Workbook()
    create_dashboard(wb, df)
    create_price_rating_sheet(wb, df)
    create_trend_sheet(wb, df)
    create_keyword_sheet(wb, df)
    create_data_sheet(wb, df)

    wb.save(OUTPUT_PATH)
    print(f"대시보드 생성 완료: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
